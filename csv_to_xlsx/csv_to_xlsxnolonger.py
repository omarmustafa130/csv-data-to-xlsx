import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
from datetime import datetime
from PIL import Image, ImageTk
import os

def convert_excel_date(excel_date):
    try:
        return datetime.strptime(str(excel_date), "%b-%d-%Y").date()
    except ValueError as e:
        print(f"Date conversion error for {excel_date}: {e}")
        return None
    
def clean_name(name):
    if pd.isna(name):
        return ''
    cleaned_name = name.split(' (')[0].replace(',', '').strip()
    return cleaned_name

def convert_csv_date(csv_date_range):
    try:
        # Split the date range and take the first part
        start_date_str = csv_date_range.split(' - ')[0].strip()
        
        # Handle different date formats that might appear in the CSV
        formats_to_try = ["%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d"]
        for fmt in formats_to_try:
            try:
                return datetime.strptime(start_date_str, fmt).date()
            except ValueError:
                pass
        
        # If no format matches, raise an error
        raise ValueError(f"Date format not recognized for {start_date_str}")
    
    except ValueError as e:
        print(f"Date conversion error for {csv_date_range}: {e}")
        return None

def get_excel_column_name(col_idx):
    """Converts 1-based column index to Excel-style column name."""
    result = ''
    while col_idx:
        mod = (col_idx - 1) % 26
        result = chr(65 + mod) + result
        col_idx = (col_idx - 1) // 26
    return result

def match_and_transfer_data(csv_file_path, excel_file_path, csv_tab_name=None, excel_tab_name=None):
    update_list = []
    
    if csv_tab_name:
        csv_data = pd.read_csv(csv_file_path, sheet_name=csv_tab_name)
    else:
        csv_data = pd.read_csv(csv_file_path)
    
    csv_data.columns = csv_data.columns.str.strip()
    
    # Clean patient names in CSV
    csv_data['Patient Name (Patient Control Number) (ID)'] = csv_data['Patient Name (Patient Control Number) (ID)'].apply(clean_name)
    
    # Read Excel data
    wb = load_workbook(excel_file_path)
    
    if excel_tab_name:
        if excel_tab_name not in wb.sheetnames:
            raise ValueError(f"Sheet {excel_tab_name} not found in Excel file.")
        ws = wb[excel_tab_name]
    else:
        ws = wb.active
    
    # Extract date columns from the first row (excluding the first column which is 'ANTHEM')
    date_columns = [cell.value for cell in ws[1]][1:]
    for index, row in csv_data.iterrows():
        patient_name = row['Patient Name (Patient Control Number) (ID)']
        total_paid_amt = row['Total Paid Amt']
        service_date_range = row['Service Dates']
        
        # Split service date range
        service_dates = service_date_range.split(' - ')
        
        for service_date_str in service_dates:
            service_date = convert_csv_date(service_date_str)
            
            if service_date is None:
                continue
            
            #print(f"Processing {patient_name} with service date: {service_date}")
            list_of_names = patient_name.split(' ')
            patient_name1 = list_of_names[0] + ' ' + list_of_names[1]
            patient_name2 = list_of_names[1] + ' ' + list_of_names[0]
            # Find the row in Excel matching the patient name
            for row_idx in range(2, ws.max_row + 1):
                excel_patient_name = clean_name(ws.cell(row=row_idx, column=1).value)
                if excel_patient_name == patient_name1 or excel_patient_name == patient_name2:
                    #print(f'Excel name: {excel_patient_name}, csv name: {patient_name}')
                    # Update corresponding date columns
                    #print(service_date)
                    for col_idx, date in enumerate(date_columns, start=2):
                        date = str(date)[:10]
                        if date == str(service_date):
                            #print(f'Excel date: {date}, csv name: {service_date}')
                            ws.cell(row=row_idx, column=col_idx, value=total_paid_amt)
                            excel_col_name = get_excel_column_name(col_idx)

                            update_list.append(f"Updated {patient_name} for date {service_date} with {total_paid_amt} at row {row_idx}, col {excel_col_name}\n")
                            break
                    break
                else:
                    pass
                    #print(f"Skipped row: Excel Name: {excel_patient_name}, CSV Name: {patient_name}")

    wb.save(excel_file_path)
    with open(os.path.join(os.getcwd(), 'updates.txt'), 'w') as f:
        for file in update_list:
            f.write(file + '\n')

def select_csv_file():
    file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    csv_file_entry.delete(0, tk.END)
    csv_file_entry.insert(0, file_path)

def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, file_path)

def toggle_tab_fields():
    global csv_label
    if use_tabs_var.get():
        csv_label.grid(row=3, column=0, padx=50, pady=20)
        csv_file_entry.grid(row=3, column=1, padx=10, pady=20)
        csv_browse.grid(row=3, column=2, padx=10, pady=10)

        excel_label.grid(row=4, column=0, padx=10, pady=20)
        excel_file_entry.grid(row=4, column=1, padx=10, pady=10)
        excel_browse.grid(row=4, column=2, padx=10, pady=10)
        

        use_tabs_check.configure(bg_color='#e8f4ff')
        run_button.grid(row=8, column=0, padx=320,pady=5)

        use_tabs_check.grid(row=9, column=0, columnspan=3, padx=50, pady=5)
        csv_tab_name_label.grid(row=5, column=0, padx=50, pady=20)
        csv_tab_name_entry.grid(row=5, column=1, padx=10, pady=20)
        excel_tab_name_label.grid(row=6, column=0, padx=50, pady=20)
        excel_tab_name_entry.grid(row=6, column=1, padx=10, pady=20)
    else:
        csv_label.grid(row=5, column=1, padx=50, pady=50)
        csv_file_entry.grid(row=5, column=2, padx=10, pady=20)
        csv_browse.grid(row=5, column=3, padx=10, pady=10)

        excel_label.grid(row=6, column=1, padx=10, pady=20)
        excel_file_entry.grid(row=6, column=2, padx=10, pady=10)
        excel_browse.grid(row=6, column=3, padx=10, pady=10)
        run_button.grid(row=8, column=1, columnspan=3, padx=270,pady=5)
        use_tabs_check.grid(row=9, column=1, columnspan=3, padx=50, pady=5)

        use_tabs_check.configure(bg_color='#f8fdff')
        csv_tab_name_label.grid_forget()
        csv_tab_name_entry.grid_forget()
        excel_tab_name_label.grid_forget()
        excel_tab_name_entry.grid_forget()

def run_program():
    csv_file_path = csv_file_entry.get()
    excel_file_path = excel_file_entry.get()
    
    if not csv_file_path or not excel_file_path:
        messagebox.showerror("Error", "Please select both CSV and Excel files.")
        return
    
    csv_tab_name = csv_tab_name_entry.get() if use_tabs_var.get() else None
    excel_tab_name = excel_tab_name_entry.get() if use_tabs_var.get() else None
    
    try:
        match_and_transfer_data(csv_file_path, excel_file_path, csv_tab_name, excel_tab_name)
        messagebox.showinfo("Success", "Data transfer completed successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

# GUI setup
root = ctk.CTk()
root.title("Data Transfer Application")
root.geometry("800x450")  # Set window size
root.resizable(False, False)  # Lock interface size

# Load and display background image
background_image = Image.open("background.png")
bg_image = ImageTk.PhotoImage(background_image)
background_label = tk.Label(root, image=bg_image)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

# Define font style
font_style = ("Inter", 15, "bold")

ctk.CTkLabel(root, text="").grid(row=0, column=0, padx=30, pady=0)
ctk.CTkLabel(root, text="").grid(row=1, column=0, padx=30, pady=0)
ctk.CTkLabel(root, text="").grid(row=2, column=0, padx=30, pady=0)

csv_label = ctk.CTkLabel(root, text="CSV File:", bg_color='#f8fdff', font=font_style)
csv_file_entry = ctk.CTkEntry(root, width=250)
csv_browse = ctk.CTkButton(root, text="Browse", command=select_csv_file, fg_color="#0086d0", font=font_style)

excel_label = ctk.CTkLabel(root, text="Excel File:", bg_color='#f8fdff', font=font_style)
excel_file_entry = ctk.CTkEntry(root, width=250)
excel_browse = ctk.CTkButton(root, text="Browse", command=select_excel_file, fg_color="#0086d0", font=font_style)

use_tabs_var = tk.BooleanVar()
use_tabs_check = ctk.CTkCheckBox(root, text="Use specific tab names", variable=use_tabs_var, command=toggle_tab_fields, bg_color='#f8fdff')
use_tabs_check.grid(row=9, column=1, columnspan=3, padx=50, pady=5)

csv_tab_name_label = ctk.CTkLabel(root, text="CSV Tab Name:", bg_color='#f8fdff', font=font_style)
csv_tab_name_entry = ctk.CTkEntry(root, width=250)

excel_tab_name_label = ctk.CTkLabel(root, text="Excel Tab Name:", bg_color='#f8fdff', font=font_style)
excel_tab_name_entry = ctk.CTkEntry(root, width=250)
csv_label.grid(row=5, column=1, padx=50, pady=50)
csv_file_entry.grid(row=5, column=2, padx=10, pady=20)
csv_browse.grid(row=5, column=3, padx=10, pady=10)

excel_label.grid(row=6, column=1, padx=10, pady=20)
excel_file_entry.grid(row=6, column=2, padx=10, pady=10)
excel_browse.grid(row=6, column=3, padx=10, pady=10)
run_button = ctk.CTkButton(root, text="Run", command=run_program, fg_color="#0086d0", font=font_style)
run_button.grid(row=8, column=1, columnspan=3, padx=270,pady=5)
root.mainloop() 
