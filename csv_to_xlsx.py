import pandas as pd
from openpyxl import load_workbook
import pdfplumber
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
from datetime import datetime
import os
import csv
import string

original_alphabet = list(string.ascii_lowercase)

def convert_excel_date(excel_date):
    try:
        return datetime.strptime(str(excel_date), "%b-%d-%Y").date()
    except ValueError as e:
        print(f"Date conversion error for {excel_date}: {e}")
        return None

def clean_name(name):
    if pd.isna(name):
        return ''
    cleaned_name = name.split(' (')[0].replace(',', '').strip()
    return cleaned_name

def convert_csv_date(csv_date_range):
    try:
        start_date_str = csv_date_range.split(' - ')[0].strip()
        formats_to_try = ["%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d"]
        for fmt in formats_to_try:
            try:
                return datetime.strptime(start_date_str, fmt).date()
            except ValueError:
                pass
        raise ValueError(f"Date format not recognized for {start_date_str}")
    except ValueError as e:
        print(f"Date conversion error for {csv_date_range}: {e}")
        return None

def normalize_spaces(s):
    """Normalize multiple spaces into a single space."""
    return ' '.join(s.split())

def match_and_transfer_data(csv_file_path, excel_file_path, tab_name=None):
    update_list = []

    csv_data = pd.read_csv(csv_file_path)
    csv_data.columns = csv_data.columns.str.strip()
    csv_data['Patient Name (Patient Control Number) (ID)'] = csv_data['Patient Name (Patient Control Number) (ID)'].apply(clean_name)
    
    wb = load_workbook(excel_file_path)

    if tab_name and tab_name in wb.sheetnames:
        ws = wb[tab_name]
    else:
        ws = wb.active

    date_columns = [cell.value for cell in ws[1]][1:]
    
    for _, row in csv_data.iterrows():
        patient_name = row['Patient Name (Patient Control Number) (ID)']
        total_paid_amt = row['Total Paid Amt']
        service_date_range = row['Service Dates']
        if not patient_name or not service_date_range or not total_paid_amt:
            print(f"Skipping row due to missing data: {row}")
            continue
        service_dates = service_date_range.split(' - ')

        for service_date_str in service_dates:
            service_date = convert_csv_date(service_date_str)
            if service_date is None:
                continue

            patient_name = normalize_spaces(patient_name)
            list_of_names = patient_name.split(' ')
            patient_name1 = ' '.join(list_of_names[:2])  # First Last
            patient_name2 = ' '.join(list_of_names[::-1])  # Last First

            for row_idx in range(2, ws.max_row + 1):
                excel_patient_name = clean_name(ws.cell(row=row_idx, column=1).value)
                excel_patient_name = normalize_spaces(excel_patient_name)
                if excel_patient_name in [patient_name1, patient_name2]:
                    for col_idx, date in enumerate(date_columns, start=2):
                        if str(date)[:10] == str(service_date):
                            ws.cell(row=row_idx, column=col_idx, value=total_paid_amt)
                            update_list.append(
                                f"Updated {patient_name} for date {service_date} with {total_paid_amt} at row {row_idx}, col {col_idx}\n"
                            )
                            break
                    break

    wb.save(excel_file_path)
    with open(os.path.join(os.getcwd(), 'updates.txt'), 'w') as f:
        for entry in update_list:
            f.write(entry)

def normalize_spaces(s):
    """Normalize multiple spaces into a single space."""
    return ' '.join(s.split())


def extract_pdf_to_csv(pdf_file_path, output_csv_path):
    """
    Extract data from PDF and save it to a CSV file.
    """
    relevant_header = [
        "LineCtrlNmbr", "Datesof Service", "Rend ProvID", "Rev",
        "SubProc/ Modifier/ Units", "AdjudProc/ Modifier/Units",
        "Remark/ PayerCode", "SuppInfo(AMT)", "Charge",
        "Adjustments (Qty)", "Adj Amount", "Payment"
    ]
    extracted_data = []

    def normalize_header(header):
        return [col.replace("\n", " ").strip() for col in header]

    with pdfplumber.open(pdf_file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            patient_name = None
            if text:
                for line in text.split('\n'):
                    if line.startswith("PatientName:"):
                        patient_name = line.split(" ")[0]
                        patient_name = patient_name.split('PatientName:')[1].split(',')
                        patient_name = patient_name[0]+' ' + patient_name[1]
                        break  # Only take the first patient name per page

            for table in page.extract_tables():
                if len(table) > 0:
                    raw_header = table[0]
                    normalized_header = normalize_header(raw_header)
                    if normalized_header == relevant_header and patient_name:
                        for row in table[1:]:
                            extracted_data.append({
                                "Patient Name": patient_name,
                                "Dates of Service": row[1].split("-")[0].strip(),
                                "Adjud Proc / Modifier / Units": row[5].split("/")[0].strip(),
                                "Charge": row[8].strip(),
                            })

    with open(output_csv_path, mode="w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(
            csv_file, fieldnames=["Patient Name", "Dates of Service", "Adjud Proc / Modifier / Units", "Charge"]
        )
        writer.writeheader()
        writer.writerows(extracted_data)

    print(f"CSV saved at {output_csv_path}")

def match_pdf_csv_to_xlsx(pdf_csv_path, excel_file_path):
    """
    Match data from a PDF-generated CSV to an Excel file, considering keywords for sheet names and years.
    """
    update_list = []
    csv_data = pd.read_csv(pdf_csv_path)

    wb = load_workbook(excel_file_path)

    # Helper function to find the closest matching sheet name
    def find_matching_sheet(keyword, year):
        for sheet_name in wb.sheetnames:
            if keyword.lower() in sheet_name.lower() and str(year) in sheet_name:
                return sheet_name
        return None  # If no matching sheet is found

    for _, row in csv_data.iterrows():
        patient_name = str(row["Patient Name"]).lower()
        dates_of_service = row["Dates of Service"]
        charge = row["Charge"]
        adjud_proc = row["Adjud Proc / Modifier / Units"]

        try:
            unformatted_date_from_csv = str(dates_of_service).split("/")
            if len(unformatted_date_from_csv) != 3:
                raise ValueError(f"Invalid date format: {dates_of_service}")
            formatted_date_from_csv = unformatted_date_from_csv[0] + "-" + unformatted_date_from_csv[1] + "-" + unformatted_date_from_csv[2]
            year_from_csv = unformatted_date_from_csv[2]
        except ValueError as e:
            print(f"Error processing date {dates_of_service}: {e}")
            continue

        # Determine the category keyword based on Adjud Proc
        if "HC:T1005" in adjud_proc:
            category_keyword = "RESPITE"
        elif "HC:T1019" in adjud_proc:
            category_keyword = "PERSONAL"
        else:
            print(f"Unknown category for Adjud Proc: {adjud_proc}. Skipping entry.")
            continue

        # Find the matching sheet name based on the category and year
        sheet_name = find_matching_sheet(category_keyword, year_from_csv)
        if not sheet_name:
            print(f"No sheet found containing '{category_keyword}' and year '{year_from_csv}'. Skipping entry.")
            continue

        ws = wb[sheet_name]
        date_columns = [cell.value for cell in ws[1]][1:]  # Get column headers from the first row
        if not date_columns:
            print(f"No date columns found in sheet '{sheet_name}'. Skipping entry.")
            continue
        # Match patient name and date in Excel
        for row_idx in range(2, ws.max_row + 1):
            excel_patient_name = str(ws.cell(row=row_idx, column=1).value).lower()
            patient_name_alphabet = [0 for _ in range(len(original_alphabet))]
            excel_patient_name_alphabet = [0 for _ in range(len(original_alphabet))]

            for letter in patient_name:
                if letter in original_alphabet:
                    index = original_alphabet.index(letter)
                    patient_name_alphabet[index] += 1
            for letter in excel_patient_name:
                if letter in original_alphabet:
                    index = original_alphabet.index(letter)
                    excel_patient_name_alphabet[index] += 1

            count_difference = 0
            same_name = True
            for i in range(len(excel_patient_name_alphabet)):
                if count_difference > 2:
                    same_name = False
                    break
                if patient_name_alphabet[i] != excel_patient_name_alphabet[i]:
                    count_difference += 1

            if same_name:
                for col_idx, date in enumerate(date_columns, start=2):
                    if date:
                        unformatted_date_from_xlsx = str(date)[:10].split("-")
                        formatted_date_from_xlsx = (
                            unformatted_date_from_xlsx[1] + "-" + unformatted_date_from_xlsx[2] + "-" + unformatted_date_from_xlsx[0]
                        )
                        if formatted_date_from_xlsx == formatted_date_from_csv:
                            ws.cell(row=row_idx, column=col_idx, value=charge)
                            update_list.append(
                                f"Updated {patient_name} for date {dates_of_service} with {charge} in sheet '{sheet_name}' at row {row_idx}, col {col_idx}\n"
                            )
                            print(
                                f"Updated {patient_name} for date {dates_of_service} with {charge} in sheet '{sheet_name}' at row {row_idx}, col {col_idx}\n"
                            )
                            break
                break

    wb.save(excel_file_path)

    # Log updates to a file
    with open(os.path.join(os.getcwd(), "updates.txt"), "w") as f:
        for entry in update_list:
            f.write(entry)

    print("Excel file updated successfully!")



def select_file(entry_field, file_types):
    file_path = filedialog.askopenfilename(filetypes=file_types)
    entry_field.delete(0, tk.END)
    entry_field.insert(0, file_path)

def toggle_file_fields(*args):
    """Dynamically adjust file fields based on the selected operation."""
    if file_type_choice.get() == "PDF to XLSX":
        pdf_file_label.grid(row=1, column=0, padx=20, pady=10)
        pdf_file_entry.grid(row=1, column=1, padx=10, pady=10)
        pdf_file_button.grid(row=1, column=2, padx=10, pady=10)
        csv_file_label.grid_forget()
        csv_file_entry.grid_forget()
        csv_file_button.grid_forget()
        tab_name_label.grid_forget()
        tab_name_entry.grid_forget()
    else:
        csv_file_label.grid(row=1, column=0, padx=20, pady=10)
        csv_file_entry.grid(row=1, column=1, padx=10, pady=10)
        csv_file_button.grid(row=1, column=2, padx=10, pady=10)
        pdf_file_label.grid_forget()
        pdf_file_entry.grid_forget()
        pdf_file_button.grid_forget()
        tab_name_label.grid(row=4, column=0, padx=20, pady=10)
        tab_name_entry.grid(row=4, column=1, padx=10, pady=10)

def run_program():
    choice = file_type_choice.get()
    excel_file_path = excel_file_entry.get()
    tab_name = tab_name_entry.get() if tab_name_entry.get() else None

    if not excel_file_path:
        messagebox.showerror("Error", "Please select an Excel file.")
        return

    try:
        if choice == "CSV to XLSX":
            csv_file_path = csv_file_entry.get()
            if not csv_file_path:
                messagebox.showerror("Error", "Please select a CSV file.")
                return
            match_and_transfer_data(csv_file_path, excel_file_path, tab_name)
        elif choice == "PDF to XLSX":
            pdf_file_path = pdf_file_entry.get()
            if not pdf_file_path:
                messagebox.showerror("Error", "Please select a PDF file.")
                return

            # Convert PDF to CSV
            csv_output_path = "pdf_output.csv"
            extract_pdf_to_csv(pdf_file_path, csv_output_path)

            # Match CSV data to Excel
            match_pdf_csv_to_xlsx(csv_output_path, excel_file_path)

            # Delete the temporary CSV file
            if os.path.exists(csv_output_path):
                os.remove(csv_output_path)
                print(f"Temporary file {csv_output_path} deleted.")

        messagebox.showinfo("Success", f"{choice} completed successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


# GUI Setup
root = ctk.CTk()
root.title("Data Transfer Application")
root.geometry("800x200")

# Disable resizing
root.resizable(False, False)
font_style = ("Inter", 15, "bold")

file_type_choice = tk.StringVar(value="CSV to XLSX")
file_type_choice.trace("w", toggle_file_fields)
ctk.CTkLabel(root, text="Select Operation:", font=font_style).grid(row=0, column=0, padx=20, pady=10)
ctk.CTkOptionMenu(root, variable=file_type_choice, values=["CSV to XLSX", "PDF to XLSX"]).grid(row=0, column=1, padx=10, pady=10)

# CSV file fields
csv_file_label = ctk.CTkLabel(root, text="CSV File:", font=font_style)
csv_file_entry = ctk.CTkEntry(root, width=400)
csv_file_button = ctk.CTkButton(root, text="Browse", command=lambda: select_file(csv_file_entry, [("CSV files", "*.csv")]))

# PDF file fields
pdf_file_label = ctk.CTkLabel(root, text="PDF File:", font=font_style)
pdf_file_entry = ctk.CTkEntry(root, width=400)
pdf_file_button = ctk.CTkButton(root, text="Browse", command=lambda: select_file(pdf_file_entry, [("PDF files", "*.pdf")]))

# Excel file fields
ctk.CTkLabel(root, text="Excel File:", font=font_style).grid(row=3, column=0, padx=20, pady=10)
excel_file_entry = ctk.CTkEntry(root, width=400)
excel_file_entry.grid(row=3, column=1, padx=10, pady=10)
ctk.CTkButton(root, text="Browse", command=lambda: select_file(excel_file_entry, [("Excel files", "*.xlsx")])).grid(row=3, column=2, padx=10, pady=10)

# Tab name field
tab_name_label = ctk.CTkLabel(root, text="Tab Name (Optional):", font=font_style)
tab_name_entry = ctk.CTkEntry(root, width=400)

# Run button
ctk.CTkButton(root, text="Run", command=run_program).grid(row=5, column=1, pady=20)

# Initialize fields
toggle_file_fields()

root.mainloop()